# Módulos de la biblioteca estándar de Python
from datetime import datetime
import json
import os
import urllib.parse
import logging
import base64
from io import BytesIO

# Módulos de bibliotecas de terceros
import openpyxl
import pytz
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, inch
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle

# Módulos de Django
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.views import LoginView
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import models
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.safestring import mark_safe
from django.utils.text import slugify

# Módulos locales del proyecto
from .forms import (
    ActaEntregaForm,
    AgregarStockConVencimientoForm,
    CustomUserCreationForm,
    CustomUserEditForm,
    DepartamentoForm,
    EliminarDepartamentoForm,
    LoteProductoForm,
    ModificarDepartamentoForm,
    ProductoForm,
    SearchUserForm,
    TransaccionForm,
    CategoriaForm,  # Añadido para manejar categorías
    ModificarCategoriaForm,  # Añadido para modificar categorías
    EliminarCategoriaForm,  # Añadido para deshabilitar categorías
)
from .models import (
    ActaEntrega,
    CustomUser,
    Departamento,
    Funcionario,
    LoteProducto,
    Producto,
    Responsable,
    Transaccion,
    Categoria,  # Añadido para manejar categorías dinámicas
)

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Funciones auxiliares
def limpiar_sesion_productos_salida(request):
    """Limpia la variable de sesión productos_salida si existe"""
    if 'productos_salida' in request.session:
        logger.info("Limpiando variable de sesión 'productos_salida'.")
        del request.session['productos_salida']
        request.session.modified = True

def paginar_resultados(request, objetos, items_por_pagina=20):
    """Aplica paginación a una lista de objetos"""
    paginator = Paginator(objetos, items_por_pagina)
    page = request.GET.get('page')
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    return page_obj

def paginar_resultados_dinamico(request, objetos, items_por_pagina=20):
    """
    Aplica paginación con numeración dinámica limitada
    Solo muestra un rango limitado de páginas para evitar sobrecarga visual
    """
    from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
    
    paginator = Paginator(objetos, items_por_pagina)
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Crear rango de páginas dinámico
    current_page = page_obj.number
    total_pages = paginator.num_pages
    
    # Configuración de la paginación dinámica
    max_pages_to_show = 10  # Máximo de páginas a mostrar
    side_pages = 2  # Páginas a mostrar a cada lado de la actual
    
    if total_pages <= max_pages_to_show:
        # Si hay pocas páginas, mostrar todas
        page_range = list(range(1, total_pages + 1))
        show_first_last = False
    else:
        # Lógica para páginas limitadas
        start_page = max(1, current_page - side_pages)
        end_page = min(total_pages, current_page + side_pages)
        
        # Ajustar para mantener siempre el mismo número de páginas visibles
        if end_page - start_page < 2 * side_pages:
            if start_page == 1:
                end_page = min(total_pages, start_page + 2 * side_pages)
            elif end_page == total_pages:
                start_page = max(1, end_page - 2 * side_pages)
        
        page_range = list(range(start_page, end_page + 1))
        show_first_last = start_page > 1 or end_page < total_pages
    
    # Agregar información adicional al objeto page
    page_obj.dynamic_page_range = page_range
    page_obj.show_first_last = show_first_last
    page_obj.show_first_ellipsis = page_range[0] > 2 if page_range else False
    page_obj.show_last_ellipsis = page_range[-1] < total_pages - 1 if page_range else False
    
    return page_obj

def generar_pdf_acta(actas, disposition='attachment'):
    """Genera un PDF para un acta de entrega con límite de 100 caracteres y texto ajustado."""
    try:
        logger.info("Generando PDF para las actas...")
        acta = actas.first()
        if not acta:
            raise ValueError("No se encontraron actas para generar el PDF.")

        productos_salida = [
            {
                'codigo_barra': item.producto.codigo_barra,
                'descripcion': item.producto.descripcion[:100],  # Límite de 100 caracteres
                'numero_siscom': str(item.numero_siscom or '')[:100],  # Límite de 100 caracteres
                'cantidad': item.cantidad,
                # Permitir observaciones largas, sin truncar, y respetar saltos de línea
                'observacion': str(item.observacion or ''),
            } for item in actas
        ]
        logger.info(f"Productos para el PDF (con límite de 100 caracteres): {productos_salida}")

        # Usar un buffer para generar el PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=0.75*inch, rightMargin=0.75*inch,
                                topMargin=0.5*inch, bottomMargin=0.5*inch,
                                title=f"Acta N°{acta.numero_acta}")  # Establecer el título del PDF
        styles = getSampleStyleSheet()

        # Definir estilos personalizados
        custom_styles = {
            'TitleCustom': ParagraphStyle(name='TitleCustom', fontName='Helvetica-Bold', fontSize=14, alignment=1, spaceAfter=10, leading=16),
            'NormalBold': ParagraphStyle(name='NormalBold', fontName='Helvetica-Bold', fontSize=10, spaceAfter=4, leading=12),
            'NormalCustom': ParagraphStyle(name='NormalCustom', fontName='Helvetica-Bold', fontSize=10, spaceAfter=4, leading=12),
            'ActaNumber': ParagraphStyle(name='ActaNumber', fontName='Helvetica-Bold', fontSize=20, alignment=1, textColor=colors.black, spaceAfter=0, leading=24),
            'Signature': ParagraphStyle(name='Signature', fontName='Times-Roman', fontSize=10, alignment=1, spaceBefore=10, spaceAfter=4, leading=12),
            'SignatureTitle': ParagraphStyle(name='SignatureTitle', fontName='Times-Bold', fontSize=10, alignment=1, spaceBefore=4, spaceAfter=4, leading=12),
            'SignatureCargo': ParagraphStyle(name='SignatureCargo', fontName='Times-Italic', fontSize=9, alignment=1, spaceBefore=2, spaceAfter=2, leading=11),
            'TableCell': ParagraphStyle(name='TableCell', fontName='Helvetica', fontSize=9, leading=11, wordWrap='CJK', alignment=0),  # Ajuste para envolver texto
        }

        elements = []
        # --- Encabezado profesional con logo y textos ---
        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'images', 'seremi_logo.png')
        logger.info(f"Ruta del logo: {logo_path}")
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=3*cm, height=3*cm)
            logger.info("Logo cargado correctamente.")
        else:
            logger.warning("Logo no encontrado, usando texto alternativo.")
            logo = Paragraph("Logo no encontrado", custom_styles['NormalCustom'])
        logo.hAlign = 'LEFT'

        encabezado_seremi = Paragraph("<b>SEREMI DE SALUD</b>", ParagraphStyle(name='EncabezadoSeremi', fontName='Helvetica-Bold', fontSize=16, alignment=0, textColor=colors.HexColor('#1a3c5e'), spaceAfter=0, leading=18))
        encabezado_region = Paragraph("REGIÓN DE LA ARAUCANÍA", ParagraphStyle(name='EncabezadoRegion', fontName='Helvetica', fontSize=13, alignment=0, textColor=colors.HexColor('#1a3c5e'), spaceAfter=10, leading=15))

        encabezado_tabla = Table(
            [
                [logo, Table([[encabezado_seremi], [encabezado_region]], colWidths=[10*cm])]
            ],
            colWidths=[3*cm, 13.5*cm], hAlign='LEFT'
        )
        encabezado_tabla.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),
            ('LEFTPADDING', (1, 0), (1, 0), 10),
        ]))

        acta_number = Paragraph(f"N° {acta.numero_acta}", custom_styles['ActaNumber'])
        acta_number_table = Table([[acta_number]], colWidths=[1.5*inch], rowHeights=[0.5*inch])
        acta_number_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 20),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ]))

        header_table = Table([[encabezado_tabla, acta_number_table]], colWidths=[13.5*cm, 4.5*cm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT')
        ]))

        # Asegurarse de que los textos estén codificados correctamente y truncados
        departamento_text = acta.departamento.encode('utf-8').decode('utf-8')[:100]
        responsable_text = (acta.responsable.nombre.encode('utf-8').decode('utf-8') if acta.responsable else 'No especificado')[:100]
        generador_text = (acta.generador.nombre.encode('utf-8').decode('utf-8') if acta.generador else 'No especificado')[:100]
        fecha_text = acta.fecha.strftime('%d-%m-%Y')
        logger.info(f"Textos para el PDF - Departamento: {departamento_text}, Responsable: {responsable_text}, Generador: {generador_text}, Fecha: {fecha_text}")

        elements.extend([
            header_table,
            Spacer(1, 0.2*cm),
            Paragraph("ACTA DE ENTREGA MATERIALES O INSUMOS", custom_styles['TitleCustom']),
            Spacer(1, 0.3*cm),
            Paragraph(f"En Temuco con fecha {fecha_text} se procede a realizar la entrega de los siguientes artículos a:", custom_styles['NormalCustom']),
            Spacer(1, 0.2*cm),
            Paragraph(f"Sección: {departamento_text}", custom_styles['NormalCustom']),
            Spacer(1, 0.1*cm),
            Paragraph(f"Responsable: {responsable_text}", custom_styles['NormalCustom']),
            Spacer(1, 0.5*cm),
            Paragraph("Datos de Productos", custom_styles['NormalBold']),
            Spacer(1, 0.2*cm)
        ])

        # Crear la tabla con ajuste de texto
        data = [['Descripción', 'Nro. SISCOM', 'Cantidad', 'Observación']]
        for item in productos_salida:
            descripcion_text = item['descripcion'].encode('utf-8').decode('utf-8')
            numero_siscom_text = item['numero_siscom'].encode('utf-8').decode('utf-8')
            # Permitir observaciones largas y saltos de línea reales
            observacion_text = (item['observacion'] or '-').replace('\n', '<br/>').encode('utf-8').decode('utf-8')
            data.append([
                Paragraph(descripcion_text, custom_styles['TableCell']),
                Paragraph(numero_siscom_text, custom_styles['TableCell']),
                Paragraph(str(item['cantidad']), custom_styles['TableCell']),
                Paragraph(observacion_text, ParagraphStyle(name='ObsCell', fontName='Helvetica', fontSize=9, leading=11, wordWrap='CJK', alignment=0, allowWidows=1, allowOrphans=1))
            ])
            logger.info(f"Fila de la tabla: {descripcion_text}, {numero_siscom_text}, {item['cantidad']}, {observacion_text}")

        # Ajustar anchos de columnas para evitar desbordamiento
        table = Table(data, colWidths=[2.2*inch, 1.8*inch, 0.8*inch, 2.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))

        responsable_lower = responsable_text.lower()
        cargo = ("SECRETARIA DEL DEPARTAMENTO" if 'secretaria' in responsable_lower 
                else "JEFE DEL DEPARTAMENTO" if 'jefe' in responsable_lower or 'jefatura' in responsable_lower 
                else "RESPONSABLE DEL DEPARTAMENTO")

        # Pie de firmas profesional y alineado
        firma_table = Table([
            [
                Paragraph('<b>' + generador_text + '</b>', ParagraphStyle(name='FirmaNombre', fontName='Helvetica-Bold', fontSize=11, alignment=1, textColor=colors.HexColor('#1a3c5e'))),
                Paragraph('<b>' + responsable_text + '</b>', ParagraphStyle(name='FirmaNombre', fontName='Helvetica-Bold', fontSize=11, alignment=1, textColor=colors.HexColor('#1a3c5e')))
            ],
            [
                Paragraph('<b>ENCARGADO DE BODEGA</b>', ParagraphStyle(name='FirmaTitulo', fontName='Helvetica-Bold', fontSize=10, alignment=1, textColor=colors.HexColor('#1a3c5e'))),
                Paragraph('<b>RECEPCIONA CONFORME</b>', ParagraphStyle(name='FirmaTitulo', fontName='Helvetica-Bold', fontSize=10, alignment=1, textColor=colors.HexColor('#1a3c5e')))
            ],
            [
                Paragraph('SEREMI DE SALUD ARAUCANÍA', ParagraphStyle(name='FirmaSub', fontName='Helvetica', fontSize=9, alignment=1, textColor=colors.HexColor('#64748b'))),
                Paragraph(cargo, ParagraphStyle(name='FirmaSub', fontName='Helvetica', fontSize=9, alignment=1, textColor=colors.HexColor('#64748b')))
            ],
            [
                Paragraph('<u>_____________________________</u>', ParagraphStyle(name='FirmaLinea', fontName='Helvetica', fontSize=12, alignment=1, textColor=colors.black)),
                Paragraph('<u>_____________________________</u>', ParagraphStyle(name='FirmaLinea', fontName='Helvetica', fontSize=12, alignment=1, textColor=colors.black))
            ]
        ], colWidths=[8*cm, 8*cm], hAlign='CENTER')
        firma_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))

        from reportlab.platypus import KeepTogether
        
        # Estrategia simple y efectiva: spacer adaptativo según cantidad de productos
        num_productos = len(productos_salida)
        
        # Calcular spacer inteligente que empuje las firmas hacia abajo
        # pero sin crear páginas innecesarias
        if num_productos <= 3:
            # Pocos productos: usar spacer grande para empujar hacia abajo
            spacer_firmas = Spacer(1, 6*cm)
        elif num_productos <= 8:
            # Productos medios: spacer moderado  
            spacer_firmas = Spacer(1, 3*cm)
        else:
            # Muchos productos: spacer mínimo para evitar nueva página
            spacer_firmas = Spacer(1, 0.8*cm)
        
        # KeepTogether asegura que las firmas no se corten
        firmas_completas = KeepTogether([
            spacer_firmas,
            firma_table
        ])
        
        elements.extend([table, firmas_completas])
        logger.info("Construyendo el PDF...")
        doc.build(elements)
        logger.info("PDF generado correctamente.")

        # Obtener el contenido del buffer y codificarlo en base64
        pdf_content = buffer.getvalue()
        pdf_base64 = base64.b64encode(pdf_content).decode('utf-8')
        buffer.close()

        # Devolver el PDF como base64 junto con el número del acta
        return {
            'pdf_base64': pdf_base64,
            'numero_acta': str(acta.numero_acta),
            'filename': f"Acta_N°{acta.numero_acta}.pdf"  # Mantener el nombre del archivo consistente
        }

    except Exception as e:
        logger.error(f"Error al generar el PDF: {str(e)}")
        return {'error': f"Error al generar el PDF: {str(e)}"}

def exportar_excel(request, datos, nombre_base, columnas, campos):
    """Genera y devuelve un archivo Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nombre_base
    ws.append(columnas)

    for col in range(1, len(columnas) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    for idx, item in enumerate(datos, start=2):
        fila = []
        for campo in campos:
            valor = getattr(item, campo) if hasattr(item, campo) else item.get(campo, '-')
            # Si el campo es 'categoria', mostramos el nombre de la categoría
            if campo == 'categoria' and valor:
                valor = valor.nombre if hasattr(valor, 'nombre') else str(valor)
            if campo == 'fecha' and isinstance(valor, datetime):
                valor = valor.strftime('%d-%m-%Y %H:%M')
            fila.append(valor)
        ws.append(fila)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        column_letter = col[0].column_letter
        if column_letter == 'A':
            ws.column_dimensions[column_letter].width = 20
        else:
            ws.column_dimensions[column_letter].width = max_length + 2

    fecha = datetime.now().strftime('%Y-%m-%d')
    filename = f"{nombre_base}_{fecha}.xlsx"
    encoded_filename = urllib.parse.quote(filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'
    wb.save(response)
    return response

# Vistas
@login_required
def home(request):
    """Vista para la página de inicio con métricas de stock"""
    limpiar_sesion_productos_salida(request)

    total_productos = Producto.objects.filter(stock__gt=0).count()
    
    if total_productos > 0:
        stock_bajo = Producto.objects.filter(stock__gte=1, stock__lte=10).count()
        stock_medio = Producto.objects.filter(stock__gt=10, stock__lte=50).count()
        stock_alto = Producto.objects.filter(stock__gt=50).count()

        porcentaje_bajo = (stock_bajo / total_productos * 100) if total_productos > 0 else 0
        porcentaje_medio = (stock_medio / total_productos * 100) if total_productos > 0 else 0
        porcentaje_alto = (stock_alto / total_productos * 100) if total_productos > 0 else 0

        porcentaje_bajo = round(porcentaje_bajo, 2)
        porcentaje_medio = round(porcentaje_medio, 2)
        porcentaje_alto = round(porcentaje_alto, 2)

        suma_porcentajes = porcentaje_bajo + porcentaje_medio + porcentaje_alto
        if suma_porcentajes != 100.0:
            if porcentaje_bajo >= porcentaje_medio and porcentaje_bajo >= porcentaje_alto:
                porcentaje_bajo = round(porcentaje_bajo + (100.0 - suma_porcentajes), 2)
            elif porcentaje_medio >= porcentaje_bajo and porcentaje_medio >= porcentaje_alto:
                porcentaje_medio = round(porcentaje_medio + (100.0 - suma_porcentajes), 2)
            else:
                porcentaje_alto = round(porcentaje_alto + (100.0 - suma_porcentajes), 2)
    else:
        stock_bajo = stock_medio = stock_alto = 0
        porcentaje_bajo = porcentaje_medio = porcentaje_alto = 0

    # Calcular métricas de vencimiento usando el sistema de lotes
    from datetime import date, timedelta
    hoy = date.today()
    
    # Productos con vencimiento que tienen stock
    productos_con_vencimiento = Producto.objects.filter(
        tiene_vencimiento=True,
        stock__gt=0
    ).prefetch_related('lotes')
    
    # Contar productos por estado de vencimiento usando el estado completo de lotes
    vencidos = criticos = precaucion = normal = 0
    for producto in productos_con_vencimiento:
        estado = producto.get_estado_vencimiento_completo()
        if estado == 'Vencido':
            vencidos += 1
        elif estado in ['Vence Hoy', 'Crítico']:
            criticos += 1
        elif estado == 'Precaución':
            precaucion += 1
        else:
            normal += 1
    
    total_con_vencimiento = productos_con_vencimiento.count()
    
    # Obtener productos más críticos para mostrar en el dashboard
    productos_criticos_list = []
    for producto in productos_con_vencimiento:
        estado = producto.get_estado_vencimiento_completo()
        if estado in ['Vencido', 'Vence Hoy', 'Crítico']:
            proximo_vencimiento = producto.get_proximo_vencimiento()
            dias_restantes = (proximo_vencimiento - hoy).days if proximo_vencimiento else None
            productos_criticos_list.append({
                'producto': producto,
                'estado': estado,
                'dias_restantes': dias_restantes,
                'fecha_vencimiento': proximo_vencimiento
            })
    
    # Ordenar por criticidad (vencidos primero, luego por días restantes)
    productos_criticos_list.sort(key=lambda x: (
        0 if x['dias_restantes'] is None or x['dias_restantes'] < 0 else x['dias_restantes'],
        x['producto'].descripcion
    ))
    
    # Tomar solo los 5 más críticos para el dashboard
    productos_criticos_top = productos_criticos_list[:5]

    chart_data = {
        'totalProductos': total_productos,
        'porcentajes': [porcentaje_bajo, porcentaje_medio, porcentaje_alto]
    }

    context = {
        'total_productos': total_productos,
        'stock_bajo': stock_bajo,
        'stock_medio': stock_medio,
        'stock_alto': stock_alto,
        'porcentaje_bajo': porcentaje_bajo,
        'porcentaje_medio': porcentaje_medio,
        'porcentaje_alto': porcentaje_alto,
        'chart_data_json': mark_safe(json.dumps(chart_data)),
        # Control de vencimientos con información de lotes
        'productos_con_vencimiento': total_con_vencimiento > 0,
        'productos_vencidos': vencidos,
        'productos_criticos': criticos,
        'productos_precaucion': precaucion,
        'productos_normal': normal,
        'total_con_vencimiento': total_con_vencimiento,
        'productos_criticos_top': productos_criticos_top,
    }
    return render(request, 'accounts/home.html', context)

class CustomLoginView(LoginView):
    """Vista personalizada para el inicio de sesión"""
    template_name = 'accounts/login.html'

    def post(self, request, *args, **kwargs):
        rut = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, rut=rut, password=password)
        if user is not None:
            login(request, user)
            return HttpResponseRedirect(self.get_success_url())
        messages.error(request, 'Credenciales inválidas')
        return self.render_to_response(self.get_context_data())

    def get_success_url(self):
        return '/'

@login_required
def registrar_producto(request):
    """Vista para registrar un nuevo producto"""
    limpiar_sesion_productos_salida(request)
    if not request.user.has_perm('accounts.can_edit'):
        messages.error(request, 'No tienes permiso para registrar productos.')
        return redirect('home')
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save()
            if producto.stock > 0:
                Transaccion(
                    producto=producto,
                    tipo='entrada',
                    cantidad=producto.stock,
                    fecha=datetime.now(pytz.UTC),
                    rut_proveedor=producto.rut_proveedor or '',
                    guia_despacho=producto.guia_despacho or '',
                    numero_factura=producto.numero_factura or '',
                    orden_compra=producto.orden_compra or '',
                    observacion='Stock inicial al registrar el producto'
                ).save()
            messages.success(request, 'Producto registrado con éxito.')
            return redirect('home')
        messages.error(request, 'Error al registrar el producto. Verifica los datos.')
    else:
        form = ProductoForm()
    return render(request, 'accounts/registrar_producto.html', {'form': form})

@login_required
def listar_productos(request):
    """Vista para listar productos con filtros y exportación a Excel"""
    limpiar_sesion_productos_salida(request)
    params = request.POST if request.method == 'POST' else request.GET
    query_codigo = params.get('codigo_barra', '')
    query_descripcion = params.get('descripcion', '')
    query_categoria = params.get('categoria', '')

    productos = Producto.objects.all().order_by('codigo_barra')
    if query_codigo:
        productos = productos.filter(codigo_barra=query_codigo)
    if query_descripcion:
        productos = productos.filter(descripcion__icontains=query_descripcion)
    if query_categoria and query_categoria != 'Todas':
        # Filtramos por el nombre de la categoría en el modelo Categoria
        productos = productos.filter(categoria__nombre=query_categoria)

    if request.method == 'POST' and 'exportar_excel' in request.POST:
        columnas = ['Código de Barra', 'Nombre del Producto', 'Categoría', 'Stock Actual']
        campos = ['codigo_barra', 'descripcion', 'categoria', 'stock']
        return exportar_excel(request, productos, "Productos", columnas, campos)

    # Obtener todas las categorías activas para el filtro
    categorias = Categoria.objects.filter(activo=True).order_by('nombre')
    # Crear la lista de categorías para el dropdown, incluyendo la opción "Todas"
    lista_categorias = [('', 'Todas')] + [(cat.nombre, cat.nombre) for cat in categorias]

    page_obj = paginar_resultados(request, productos)
    context = {
        'page_obj': page_obj,
        'query_codigo': query_codigo,
        'query_descripcion': query_descripcion,
        'query_categoria': query_categoria,
        'categorias': lista_categorias,  # Usamos las categorías dinámicas
    }
    return render(request, 'accounts/listar_productos.html', context)

@login_required
def agregar_stock(request):
    """Vista para listar productos y agregar stock"""
    limpiar_sesion_productos_salida(request)
    if not request.user.has_perm('accounts.can_edit'):
        messages.error(request, 'No tienes permiso para agregar stock.')
        return redirect('home')
    productos = Producto.objects.all().order_by('codigo_barra')
    query_codigo = request.GET.get('codigo_barra', '')
    query_descripcion = request.GET.get('descripcion', '')

    if query_codigo:
        productos = productos.filter(codigo_barra=query_codigo)
    if query_descripcion:
        productos = productos.filter(descripcion__icontains=query_descripcion)

    page_obj = paginar_resultados(request, productos)
    return render(request, 'accounts/agregar_stock.html', {
        'page_obj': page_obj,
        'query_codigo': query_codigo,
        'query_descripcion': query_descripcion,
    })

@login_required
def agregar_stock_detalle(request, codigo_barra):
    """Vista para agregar stock a un producto específico con manejo automático de lotes"""
    limpiar_sesion_productos_salida(request)
    if not request.user.has_perm('accounts.can_edit'):
        messages.error(request, 'No tienes permiso para agregar stock.')
        return redirect('home')
    try:
        producto = Producto.objects.get(codigo_barra=codigo_barra)
    except Producto.DoesNotExist:
        messages.error(request, 'Producto no encontrado.')
        return redirect('agregar-stock')

    if request.method == 'POST':
        form = AgregarStockConVencimientoForm(request.POST, producto=producto)
        if form.is_valid():
            try:
                # Usar el método del formulario para agregar stock con lotes automáticos
                lote, transaccion = form.agregar_stock_a_producto(producto)
                
                if lote:
                    messages.success(request, 
                        f'Stock agregado exitosamente al producto "{producto.descripcion}". '
                        f'Lote #{lote.numero_lote} - Cantidad: {form.cleaned_data["cantidad"]} unidades. '
                        f'Vence: {lote.fecha_vencimiento.strftime("%d/%m/%Y")}. '
                        f'Stock total: {producto.stock} unidades'
                    )
                else:
                    messages.success(request, 
                        f'Stock agregado exitosamente al producto "{producto.descripcion}". '
                        f'Cantidad: {form.cleaned_data["cantidad"]} unidades. '
                        f'Stock total: {producto.stock} unidades'
                    )
                
                return redirect('agregar-stock')
                
            except Exception as e:
                messages.error(request, f'Error al agregar stock: {str(e)}')
        else:
            messages.error(request, 'Error al agregar stock. Verifica los datos.')
    else:
        form = AgregarStockConVencimientoForm(producto=producto)
    
    # Obtener información de lotes existentes para mostrar en la vista
    lotes_existentes = []
    info_proximo_lote = None
    
    if producto.tiene_vencimiento:
        lotes_existentes = producto.get_lotes_detalle()
        info_proximo_lote = producto.get_info_proximo_lote()
    
    context = {
        'form': form, 
        'producto': producto,
        'tiene_vencimiento_actual': producto.tiene_vencimiento,
        'lotes_existentes': lotes_existentes,
        'total_lotes': len(lotes_existentes),
        'info_proximo_lote': info_proximo_lote
    }
    return render(request, 'accounts/agregar_stock_detalle.html', context)

@login_required
def salida_productos(request):
    """Vista para gestionar la salida de productos"""
    if not request.user.has_perm('accounts.can_edit'):
        messages.error(request, 'No tienes permiso para realizar salidas de productos.')
        return redirect('home')

    # Inicializar o cargar productos_salida desde la sesión
    if 'productos_salida' not in request.session:
        logger.info("Inicializando productos_salida en la sesión como lista vacía.")
        request.session['productos_salida'] = []
        request.session.modified = True  # Forzar sincronización de la sesión

    productos_salida = request.session.get('productos_salida', [])
    logger.info(f"Contenido inicial de productos_salida en la sesión: {productos_salida}")

    # Preparar la lista de productos para mostrar
    productos = Producto.objects.all().order_by('codigo_barra')
    query_codigo = request.GET.get('codigo_barra', '')
    query_descripcion = request.GET.get('descripcion', '')

    if query_codigo:
        productos = productos.filter(codigo_barra=query_codigo)
    if query_descripcion:
        productos = productos.filter(descripcion__icontains=query_descripcion)

    page_obj = paginar_resultados(request, productos)

    # Manejar solicitud AJAX para obtener datos de salida
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.GET.get('action') == 'get_salida_data':
        logger.info("Solicitud AJAX para obtener datos de salida.")
        productos_salida = request.session.get('productos_salida', [])  # Volver a cargar desde la sesión
        logger.info(f"Enviando productos_salida en respuesta AJAX: {productos_salida}")
        return JsonResponse({'success': True, 'productos_salida': productos_salida})

    if request.method == 'POST':
        logger.info(f"POST recibido en salida_productos: {request.POST}")

        # Manejar solicitudes AJAX
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Acción para agregar un producto
            if 'agregar_producto' in request.POST:
                codigo_barra = request.POST.get('codigo_barra')
                logger.info(f"Intentando agregar producto con código de barra: {codigo_barra}")
                try:
                    producto = Producto.objects.get(codigo_barra=codigo_barra)
                    productos_salida = request.session.get('productos_salida', [])  # Recargar desde la sesión
                    logger.info(f"Lista actual de productos_salida antes de agregar: {productos_salida}")

                    if any(item['codigo_barra'] == codigo_barra for item in productos_salida):
                        logger.warning(f"El producto {codigo_barra} ya está en la lista de salida.")
                        return JsonResponse({'success': False, 'error': 'Este producto ya está en la lista de salida.'})
                    if producto.stock == 0:
                        logger.warning(f"No se puede retirar el producto {codigo_barra} porque no tiene stock.")
                        return JsonResponse({'success': False, 'error': 'No se puede retirar este producto porque no tiene stock.'})

                    productos_salida.append({
                        'codigo_barra': producto.codigo_barra,
                        'descripcion': producto.descripcion,
                        'stock': producto.stock,
                        'numero_siscom': '',
                        'cantidad': '',  # Inicialmente vacío
                        'observacion': '',
                    })
                    request.session['productos_salida'] = productos_salida
                    request.session.modified = True
                    logger.info(f"Producto agregado a la lista de salida: {producto.codigo_barra}")
                    logger.info(f"Lista actualizada de productos_salida: {productos_salida}")
                    return JsonResponse({'success': True})

                except Producto.DoesNotExist:
                    logger.error(f"Producto con código {codigo_barra} no encontrado.")
                    return JsonResponse({'success': False, 'error': 'Producto no encontrado.'})

            # Acción para eliminar un producto
            elif 'eliminar_producto' in request.POST:
                codigo_barra = request.POST.get('codigo_barra')
                logger.info(f"Intentando eliminar producto con código de barra: {codigo_barra}")
                productos_salida = request.session.get('productos_salida', [])  # Recargar desde la sesión
                logger.info(f"Lista actual de productos_salida antes de eliminar: {productos_salida}")
                productos_salida = [item for item in productos_salida if item['codigo_barra'] != codigo_barra]
                request.session['productos_salida'] = productos_salida
                request.session.modified = True
                logger.info(f"Producto eliminado de la lista de salida: {codigo_barra}")
                logger.info(f"Lista actualizada de productos_salida: {productos_salida}")
                return JsonResponse({'success': True})

            # Acción para actualizar datos de un producto
            elif request.POST.get('action') == 'update_data':
                codigo_barra = request.POST.get('codigo_barra')
                numero_siscom = request.POST.get('numero_siscom', '').strip()
                cantidad = request.POST.get('cantidad', '').strip()
                observacion = request.POST.get('observacion', '').strip()

                logger.info(f"Actualizando datos del producto {codigo_barra}: SISCOM={numero_siscom}, Cantidad={cantidad}, Observación={observacion}")

                try:
                    producto = Producto.objects.get(codigo_barra=codigo_barra)
                    productos_salida = request.session.get('productos_salida', [])  # Recargar desde la sesión
                    logger.info(f"Lista actual de productos_salida antes de actualizar: {productos_salida}")

                    # Validar número SISCOM
                    if numero_siscom and not numero_siscom.isdigit():
                        logger.warning(f"Número SISCOM inválido para el producto {codigo_barra}: {numero_siscom}")
                        return JsonResponse({'success': False, 'error': f'El Número de SISCOM para el producto {codigo_barra} debe ser un número entero.'})

                    # Validar cantidad solo si no está vacía
                    if cantidad:
                        if not cantidad.isdigit():
                            logger.warning(f"Cantidad inválida para el producto {codigo_barra}: {cantidad}")
                            return JsonResponse({'success': False, 'error': f'La cantidad para el producto {codigo_barra} debe ser un número entero.'})
                        cantidad_int = int(cantidad)
                        if cantidad_int <= 0:
                            logger.warning(f"Cantidad no positiva para el producto {codigo_barra}: {cantidad_int}")
                            return JsonResponse({'success': False, 'error': f'La cantidad para el producto {codigo_barra} debe ser mayor que 0.'})
                        if cantidad_int > producto.stock:
                            logger.warning(f"Cantidad excede el stock para el producto {codigo_barra}. Cantidad: {cantidad_int}, Stock: {producto.stock}")
                            return JsonResponse({'success': False, 'error': f'La cantidad para el producto {codigo_barra} no puede superar el stock ({producto.stock}).'})

                    # Actualizar el producto en la lista
                    updated = False
                    for item in productos_salida:
                        if item['codigo_barra'] == codigo_barra:
                            item.update({
                                'numero_siscom': numero_siscom,
                                'cantidad': cantidad,  # Mantener el valor de cantidad
                                'observacion': observacion,
                                'stock': producto.stock,
                            })
                            updated = True
                            logger.info(f"Datos actualizados para el producto {codigo_barra}: {item}")
                            break

                    if updated:
                        request.session['productos_salida'] = productos_salida
                        request.session.modified = True
                        logger.info(f"Lista actualizada de productos_salida después de actualizar: {productos_salida}")
                        return JsonResponse({'success': True})
                    else:
                        logger.warning(f"No se encontró el producto {codigo_barra} en la lista de salida.")
                        return JsonResponse({'success': False, 'error': f'No se encontró el producto {codigo_barra} en la lista de salida.'})

                except Producto.DoesNotExist:
                    logger.error(f"Producto con código {codigo_barra} no encontrado.")
                    return JsonResponse({'success': False, 'error': 'Producto no encontrado.'})

        # Manejar el formulario de salida de productos (botón "Siguiente")
        if 'siguiente' in request.POST:
            logger.info("Procesando formulario de salida de productos (botón 'Siguiente')")
            productos_salida = request.session.get('productos_salida', [])  # Recargar desde la sesión
            logger.info(f"Productos en la sesión antes de validar: {productos_salida}")

            # Verificar si hay productos en la lista de salida
            if not productos_salida:
                logger.warning("No hay productos en la lista de salida.")
                messages.error(request, 'Debes agregar al menos un producto para continuar.')
                return redirect('salida-productos')

            # Validar cada producto en la lista de salida
            for item in productos_salida:
                logger.info(f"Validando producto: {item}")

                # Validar número SISCOM
                numero_siscom = str(item.get('numero_siscom', '')).strip()
                if not numero_siscom:
                    logger.warning(f"Número SISCOM vacío para el producto {item['codigo_barra']}")
                    messages.error(request, f"El Número de SISCOM para el producto {item['codigo_barra']} no puede estar vacío.")
                    return redirect('salida-productos')
                if not numero_siscom.isdigit():
                    logger.warning(f"Número SISCOM inválido para el producto {item['codigo_barra']}: {numero_siscom}")
                    messages.error(request, f"El Número de SISCOM para el producto {item['codigo_barra']} debe ser un número entero válido (valor recibido: '{numero_siscom}').")
                    return redirect('salida-productos')

                # Validar cantidad
                cantidad_str = str(item.get('cantidad', '')).strip()
                if not cantidad_str:
                    logger.warning(f"Cantidad vacía para el producto {item['codigo_barra']}")
                    messages.error(request, f"La cantidad para el producto {item['codigo_barra']} no puede estar vacía.")
                    return redirect('salida-productos')

                try:
                    cantidad = int(cantidad_str)
                    logger.info(f"Cantidad convertida para el producto {item['codigo_barra']}: {cantidad}")

                    # Validar que la cantidad sea positiva
                    if cantidad <= 0:
                        logger.warning(f"Cantidad no positiva para el producto {item['codigo_barra']}: {cantidad}")
                        messages.error(request, f"La cantidad para el producto {item['codigo_barra']} debe ser mayor que 0 (valor recibido: {cantidad}).")
                        return redirect('salida-productos')

                    # Validar que la cantidad no supere el stock
                    try:
                        producto = Producto.objects.get(codigo_barra=item['codigo_barra'])
                        if cantidad > producto.stock:
                            logger.warning(f"Cantidad excede el stock para el producto {item['codigo_barra']}. Cantidad: {cantidad}, Stock: {producto.stock}")
                            messages.error(request, f"La cantidad a retirar ({cantidad}) para el producto {item['codigo_barra']} no puede superar el stock actual ({producto.stock}).")
                            return redirect('salida-productos')
                        item['stock'] = producto.stock  # Actualizar el stock en la sesión
                    except Producto.DoesNotExist:
                        logger.error(f"Producto con código {item['codigo_barra']} no encontrado durante la validación.")
                        messages.error(request, f"El producto con código {item['codigo_barra']} no existe.")
                        return redirect('salida-productos')

                    # Actualizar cantidad en el producto
                    item['cantidad'] = cantidad

                except ValueError:
                    logger.warning(f"Cantidad no convertible a entero para el producto {item['codigo_barra']}: {cantidad_str}")
                    messages.error(request, f"La cantidad para el producto {item['codigo_barra']} debe ser un número entero (valor recibido: '{cantidad_str}').")
                    return redirect('salida-productos')

            # Si todas las validaciones pasaron, redirigir a la siguiente etapa
            logger.info("Todas las validaciones pasaron. Redirigiendo a salida-productos-seleccion")
            request.session['productos_salida'] = productos_salida
            request.session.modified = True
            logger.info(f"Productos guardados en la sesión antes de redirigir: {request.session['productos_salida']}")
            return redirect('salida-productos-seleccion')

        else:
            logger.warning("Acción POST no reconocida en salida_productos.")
            messages.error(request, "Acción no reconocida. Por favor, intenta de nuevo.")

    # Manejar solicitud AJAX para renderizar la tabla de productos disponibles
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.method == 'GET':
        productos_salida = request.session.get('productos_salida', [])  # Volver a cargar desde la sesión
        logger.info(f"Renderizando tabla de productos disponibles con productos_salida: {productos_salida}")
        context = {
            'page_obj': page_obj,
            'query_codigo': query_codigo,
            'query_descripcion': query_descripcion,
            'productos_salida': productos_salida,  # Asegurarse de pasar la lista actualizada
        }
        return render(request, 'accounts/salida_productos.html', context)

    logger.info(f"Renderizando salida_productos.html con productos_salida: {productos_salida}")
    return render(request, 'accounts/salida_productos.html', {
        'page_obj': page_obj,
        'query_codigo': query_codigo,
        'query_descripcion': query_descripcion,
        'productos_salida': productos_salida,
    })

@login_required
def salida_productos_seleccion(request):
    """Vista para seleccionar departamento y responsable, y generar el acta de entrega"""
    if not request.user.has_perm('accounts.can_edit'):
        messages.error(request, 'No tienes permiso para realizar salidas de productos.')
        return redirect('home')

    # Verificar si hay productos en la sesión
    productos_salida = request.session.get('productos_salida', [])
    if not productos_salida:
        logger.warning("No hay productos seleccionados para la salida en salida_productos_seleccion.")
        messages.error(request, 'No hay productos seleccionados para la salida.')
        return redirect('salida-productos')

    # Verificar si el acta ya fue generada para evitar duplicados
    if 'acta_generada' in request.session and request.session['acta_generada']:
        logger.info("Acta ya generada, redirigiendo a salida-productos.")
        messages.info(request, 'El acta ya ha sido generada. Por favor, inicia una nueva salida.')
        return redirect('salida-productos')

    if request.method == 'POST':
        # Manejar el botón "Cancelar"
        if 'cancelar' in request.POST:
            logger.info("Botón 'Cancelar' presionado en salida_productos_seleccion. Redirigiendo a salida_productos.")
            # No limpiamos productos_salida, simplemente redirigimos
            return redirect('salida-productos')

        form = ActaEntregaForm(request.POST)
        logger.info(f"Formulario recibido en salida_productos_seleccion: {request.POST}")
        if form.is_valid():
            logger.info("Formulario válido. Procesando la salida...")
            logger.info(f"Datos limpiados - Departamento: {form.cleaned_data['departamento']}, Responsable: {form.cleaned_data['responsable']}")
            try:
                # Validar el stock disponible directamente del producto (ya sincronizado)
                for item in productos_salida:
                    logger.info(f"Validando stock para el producto: {item}")
                    producto = Producto.objects.get(codigo_barra=item['codigo_barra'])
                    cantidad = int(item['cantidad'] or 0)

                    # Usar stock del producto directamente (ya está sincronizado con lotes)
                    stock_disponible = producto.stock
                    
                    logger.info(f"Producto: {producto.descripcion}, Stock disponible: {stock_disponible}, Cantidad solicitada: {cantidad}")
                    if stock_disponible < cantidad:
                        logger.warning(f"No hay suficiente stock para {producto.descripcion} (Código: {producto.codigo_barra}). Stock disponible: {stock_disponible}, Solicitado: {cantidad}.")
                        messages.error(request, f'No hay suficiente stock para {producto.descripcion} (Código: {producto.codigo_barra}). Stock disponible: {stock_disponible}, Solicitado: {cantidad}.')
                        return redirect('salida-productos-seleccion')

                ultimo_acta = ActaEntrega.objects.order_by('-numero_acta').first()
                numero_acta = 1 if not ultimo_acta else ultimo_acta.numero_acta + 1
                logger.info(f"Nuevo número de acta: {numero_acta}")

                responsable = form.cleaned_data['responsable']

                for item in productos_salida:
                    logger.info(f"Creando acta para el producto: {item}")
                    producto = Producto.objects.get(codigo_barra=item['codigo_barra'])
                    cantidad = int(item['cantidad'])
                    acta = ActaEntrega(
                        numero_acta=numero_acta,
                        departamento=form.cleaned_data['departamento'],
                        responsable=responsable,
                        generador=request.user,
                        producto=producto,
                        cantidad=cantidad,
                        numero_siscom=item['numero_siscom'],
                        observacion=item['observacion'],
                    )
                    acta.save()
                    logger.info(f"Acta creada: N°{acta.numero_acta}, Producto: {producto.descripcion}, Cantidad: {cantidad}")

                    Transaccion.objects.create(
                        producto=producto,
                        tipo='salida',
                        cantidad=cantidad,
                        acta_entrega=acta,
                        fecha=datetime.now(pytz.UTC),
                        observacion=f"Salida asociada al Acta N°{numero_acta}"
                    )
                    logger.info(f"Transacción creada: Tipo: salida, Cantidad: {cantidad}")

                    # Usar sistema FIFO automático para reducir stock
                    if producto.tiene_vencimiento:
                        exito = producto.reducir_stock_fifo(cantidad)
                        if not exito:
                            logger.error(f"Error al reducir stock FIFO para {producto.descripcion}")
                            messages.error(request, f'Error al reducir stock para {producto.descripcion}')
                            return JsonResponse({'success': False, 'error': f'Error al reducir stock para {producto.descripcion}'})
                        logger.info(f"Stock reducido usando FIFO: {producto.descripcion}, Nuevo stock: {producto.stock}")
                    else:
                        # Para productos sin vencimiento, reducir del stock principal
                        producto.stock -= cantidad
                        producto.save()
                        logger.info(f"Stock reducido directamente: {producto.descripcion}, Nuevo stock: {producto.stock}")

                actas = ActaEntrega.objects.filter(numero_acta=numero_acta)
                logger.info(f"Actas para el PDF: {list(actas)}")

                # Marcar el acta como generada para evitar duplicados
                request.session['acta_generada'] = True
                request.session.modified = True

                # Generar el PDF y obtener el resultado
                pdf_result = generar_pdf_acta(actas)

                # Verificar si hubo un error al generar el PDF
                if 'error' in pdf_result:
                    logger.error(f"Error al generar el PDF: {pdf_result['error']}")
                    messages.error(request, pdf_result['error'])
                    return redirect('salida-productos-seleccion')

                # Preparar la respuesta JSON
                response_data = {
                    'success': True,
                    'pdf_base64': pdf_result['pdf_base64'],
                    'numero_acta': pdf_result['numero_acta'],
                    'filename': pdf_result['filename'],
                    'message': f'Acta de entrega N°{numero_acta} generada correctamente.'
                }

                # Limpiar la sesión después de generar el PDF
                limpiar_sesion_productos_salida(request)
                if 'acta_generada' in request.session:
                    del request.session['acta_generada']
                    request.session.modified = True

                return JsonResponse(response_data)

            except Exception as e:
                logger.error(f"Error al procesar el acta: {str(e)}")
                messages.error(request, f'Error al procesar el acta: {str(e)}')
                return JsonResponse({'success': False, 'error': f'Error al procesar el acta: {str(e)}'})
        else:
            logger.warning("Formulario no válido en salida_productos_seleccion.")
            errors = {}
            for field, error_list in form.errors.items():
                errors[field] = [str(error) for error in error_list]
            return JsonResponse({'success': False, 'errors': errors})
    else:
        form = ActaEntregaForm()
        # Asegurar que la bandera de acta generada esté inicializada
        request.session['acta_generada'] = False
        request.session.modified = True

    return render(request, 'accounts/salida_productos_seleccion.html', {'form': form, 'productos_salida': productos_salida})

@login_required
def listar_actas(request):
    """Vista para listar las actas de entrega"""
    limpiar_sesion_productos_salida(request)
    actas = ActaEntrega.objects.all().order_by('-numero_acta')
    query_numero_acta = request.GET.get('numero_acta', '')
    query_responsable = request.GET.get('responsable', '')

    if query_numero_acta:
        try:
            actas = actas.filter(numero_acta__startswith=int(query_numero_acta))
        except ValueError:
            messages.error(request, 'El número de acta debe ser un valor numérico.')
    if query_responsable:
        actas = actas.filter(responsable__nombre__icontains=query_responsable)

    actas_dict = {acta.numero_acta: acta for acta in actas}
    actas_lista = sorted(actas_dict.values(), key=lambda x: x.numero_acta, reverse=True)

    page_obj = paginar_resultados(request, actas_lista, items_por_pagina=20)

    return render(request, 'accounts/listar_actas.html', {
        'actas': page_obj,
        'query_numero_acta': query_numero_acta,
        'query_responsable': query_responsable,
    })

@login_required
def ver_acta_pdf(request, numero_acta, disposition):
    """Vista para visualizar un acta de entrega en PDF"""
    actas = ActaEntrega.objects.filter(numero_acta=numero_acta)
    if not actas.exists():
        return HttpResponse("Acta no encontrada.", status=404)

    pdf_result = generar_pdf_acta(actas, disposition)
    if 'error' in pdf_result:
        return HttpResponse(pdf_result['error'], status=500)

    # Crear una respuesta HTTP para el PDF
    response = HttpResponse(content_type='application/pdf')
    filename = f"Acta_N°{numero_acta}.pdf"  # Cambiado para incluir "Acta_N°" en el nombre
    encoded_filename = urllib.parse.quote(filename)
    response['Content-Disposition'] = f'{disposition}; filename="{encoded_filename}"'
    response['Content-Type'] = 'application/pdf; charset=utf-8'

    # Decodificar el PDF desde base64 y escribirlo en la respuesta
    pdf_content = base64.b64decode(pdf_result['pdf_base64'])
    response.write(pdf_content)
    return response

@login_required
def bincard_buscar(request):
    """Vista para buscar un producto por código de barra y ver su historial"""
    limpiar_sesion_productos_salida(request)
    if request.method == 'POST':
        codigo_barra = request.POST.get('codigo_barra', '').strip()
        if not codigo_barra:
            messages.error(request, 'Por favor, ingrese un código de barra.')
        elif not codigo_barra.isdigit():
            messages.error(request, 'El código de barra debe contener solo números.')
        else:
            try:
                Producto.objects.get(codigo_barra=codigo_barra)
                return redirect('bincard-historial', codigo_barra=codigo_barra)
            except Producto.DoesNotExist:
                messages.error(request, f'No se encontró un producto con el código de barra {codigo_barra}. '
                                f'Puedes <a href="/accounts/registrar-producto/">registrar un nuevo producto</a>.')
        return render(request, 'accounts/bincard_buscar.html')
    return render(request, 'accounts/bincard_buscar.html')

@login_required
def buscar_codigos_barra(request):
    """Vista para buscar códigos de barra mediante autocompletado"""
    term = request.GET.get('term', '').strip()
    if not term:
        return JsonResponse([], safe=False)
    productos = Producto.objects.filter(codigo_barra__startswith=term).order_by('codigo_barra')[:10]
    codigos = [{'label': f"{p.codigo_barra} - {p.descripcion}", 'value': p.codigo_barra} for p in productos]
    return JsonResponse(codigos, safe=False)

@login_required
def bincard_historial(request, codigo_barra):
    """Vista para mostrar el historial de transacciones de un producto"""
    limpiar_sesion_productos_salida(request)
    if not codigo_barra.isdigit():
        messages.error(request, 'El código de barra debe contener solo números.')
        return redirect('bincard-buscar')

    try:
        producto = Producto.objects.get(codigo_barra=codigo_barra)
    except Producto.DoesNotExist:
        messages.error(request, 'Producto no encontrado.')
        return redirect('bincard-buscar')

    transacciones = Transaccion.objects.filter(producto=producto).order_by('fecha')
    actas = ActaEntrega.objects.filter(producto=producto).order_by('fecha')

    eventos = []
    for transaccion in transacciones.filter(tipo='entrada'):
        guia_o_factura = transaccion.guia_despacho or transaccion.numero_factura or "-"
        if transaccion.guia_despacho:
            guia_o_factura = f"Guía: {transaccion.guia_despacho}"
        elif transaccion.numero_factura:
            guia_o_factura = f"Factura: {transaccion.numero_factura}"

        eventos.append({
            'tipo': 'entrada',
            'fecha': transaccion.fecha or datetime.now(pytz.UTC),
            'guia_o_factura': guia_o_factura,
            'numero_acta': None,
            'rut_proveedor': transaccion.rut_proveedor or '-',
            'departamento': None,
            'entrada': transaccion.cantidad,
            'salida': 0,
        })

    for transaccion in transacciones.filter(tipo='salida'):
        if transaccion.acta_entrega:
            eventos.append({
                'tipo': 'salida',
                'fecha': transaccion.fecha or transaccion.acta_entrega.fecha or datetime.now(pytz.UTC),
                'guia_o_factura': "-",
                'numero_acta': transaccion.acta_entrega.numero_acta,
                'rut_proveedor': None,
                'departamento': transaccion.acta_entrega.departamento,
                'entrada': 0,
                'salida': transaccion.cantidad,
            })

    eventos.sort(key=lambda x: x['fecha'] or datetime.now(pytz.UTC))

    saldo = 0
    movimientos = []
    for evento in eventos:
        if evento['tipo'] == 'entrada':
            saldo += evento['entrada']
        else:
            saldo -= evento['salida']

        if saldo < 0:
            messages.error(request, f'Error: El saldo no puede ser negativo en la fecha {evento["fecha"]}. Contacte al administrador.')
            return redirect('bincard-buscar')

        evento['saldo'] = saldo
        movimientos.append(evento)

    total_entradas = sum(m['entrada'] for m in movimientos)
    total_salidas = sum(m['salida'] for m in movimientos)

    # CORRECCIÓN DEFINITIVA: Manejo diferenciado para productos con/sin vencimiento
    if producto.tiene_vencimiento:
        # Para productos con vencimiento: Los lotes son la fuente de verdad
        # CRÍTICO: NO limpiar lotes vacíos - preservar trazabilidad para Bincard
        stock_real_lotes = producto.lotes.aggregate(total=models.Sum('stock'))['total'] or 0
        
        # Sincronizar stock del producto con la suma de lotes (fuente de verdad)
        if stock_real_lotes != producto.stock:
            producto.stock = stock_real_lotes
            producto.save()
            messages.info(request, f'Stock sincronizado: {stock_real_lotes} unidades (desde lotes activos).')
        
        # Para productos con vencimiento, no validamos contra el saldo histórico
        # porque las operaciones FIFO pueden generar diferencias legítimas
        if abs(saldo - stock_real_lotes) > 0:
            messages.info(request, f'El saldo histórico ({saldo}) puede diferir del stock real ({stock_real_lotes}) debido al manejo FIFO de lotes con vencimiento.')
            
    else:
        # Para productos sin vencimiento: El historial lineal es la fuente de verdad
        if saldo != producto.stock:
            messages.warning(request, f'Advertencia: El saldo calculado ({saldo}) no coincide con el stock actual del producto ({producto.stock}).')
            producto.stock = saldo
            producto.save()
            messages.info(request, f'El stock del producto ha sido corregido a {saldo} para que coincida con el saldo calculado.')

    if request.method == 'POST' and 'exportar_excel' in request.POST:
        columnas = ['Fecha', 'Guía o Factura', 'N° Acta', 'Proveedor (RUT)', 'Programa/Departamento', 'Entrada', 'Salida', 'Saldo']
        campos = ['fecha', 'guia_o_factura', 'numero_acta', 'rut_proveedor', 'departamento', 'entrada', 'salida', 'saldo']
        return exportar_excel(request, movimientos, f"Bincard_{producto.codigo_barra}", columnas, campos)

    page_obj = paginar_resultados(request, movimientos)
    return render(request, 'accounts/bincard_historial.html', {
        'producto': producto,
        'page_obj': page_obj,
        'total_entradas': total_entradas,
        'total_salidas': total_salidas,
    })

@login_required
def agregar_departamento(request):
    """Vista para agregar un nuevo departamento"""
    limpiar_sesion_productos_salida(request)
    if not request.user.has_perm('accounts.can_manage_departments'):
        messages.error(request, 'No tienes permiso para agregar departamentos.')
        return redirect('home')
    if request.method == 'POST':
        form = DepartamentoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Departamento agregado con éxito.')
            return redirect('home')
        messages.error(request, 'Error al agregar el departamento. Verifica los datos.')
    else:
        form = DepartamentoForm()
    return render(request, 'accounts/agregar_departamento.html', {'form': form})

@login_required
def detalle_lotes_producto(request, codigo_barra):
    """Vista para ver el detalle de todos los lotes de un producto específico."""
    try:
        producto = Producto.objects.get(codigo_barra=codigo_barra)
    except Producto.DoesNotExist:
        messages.error(request, 'Producto no encontrado.')
        return redirect('control-vencimientos')
    
    if not producto.tiene_vencimiento:
        messages.info(request, 'Este producto no maneja lotes porque no tiene fecha de vencimiento.')
        return redirect('control-vencimientos')
    
    # Obtener todos los lotes (con y sin stock) para historial completo
    lotes_con_stock = producto.lotes.filter(stock__gt=0).order_by('fecha_vencimiento')
    lotes_sin_stock = producto.lotes.filter(stock=0).order_by('-fecha_ingreso')[:10]  # Últimos 10 lotes vacíos
    
    # Información detallada de lotes
    lotes_detalle = producto.get_lotes_detalle()
    
    # Calcular estadísticas de lotes
    total_lotes = producto.lotes.count()
    lotes_activos = lotes_con_stock.count()
    proximo_vencimiento = producto.get_proximo_vencimiento()
    estado_general = producto.get_estado_vencimiento_completo()
    
    from datetime import date
    hoy = date.today()
    
    context = {
        'producto': producto,
        'lotes_con_stock': lotes_con_stock,
        'lotes_sin_stock': lotes_sin_stock,
        'lotes_detalle': lotes_detalle,
        'total_lotes': total_lotes,
        'lotes_activos': lotes_activos,
        'proximo_vencimiento': proximo_vencimiento,
        'estado_general': estado_general,
        'hoy': hoy,
    }
    
    return render(request, 'accounts/detalle_lotes_producto.html', context)

@login_required
def modificar_departamento(request):
    """Vista para modificar un departamento existente"""
    limpiar_sesion_productos_salida(request)
    if not request.user.has_perm('accounts.can_manage_departments'):
        messages.error(request, 'No tienes permiso para modificar departamentos.')
        return redirect('home')
    if request.method == 'POST':
        form = ModificarDepartamentoForm(request.POST)
        if form.is_valid():
            departamento_nombre = form.cleaned_data['departamento']
            nuevo_nombre = form.cleaned_data['nuevo_nombre']
            jefatura = form.cleaned_data['jefatura']
            jefatura_subrogante = form.cleaned_data['jefatura_subrogante']
            secretaria = form.cleaned_data['secretaria']
            secretaria_subrogante = form.cleaned_data['secretaria_subrogante']

            departamento = Departamento.objects.get(nombre=departamento_nombre, activo=True)
            departamento.nombre = nuevo_nombre
            departamento.save()

            responsables = Responsable.objects.filter(departamento=departamento)
            
            if not jefatura:
                responsables.filter(tipo='Jefatura').update(nombre=f"Jefatura {nuevo_nombre}")
            else:
                responsables.filter(tipo='Jefatura').update(nombre=jefatura)

            if not jefatura_subrogante:
                responsables.filter(tipo='Jefatura Subrogante').update(nombre=f"Jefatura {nuevo_nombre}(s)")
            else:
                responsables.filter(tipo='Jefatura Subrogante').update(nombre=jefatura_subrogante)

            if not secretaria:
                responsables.filter(tipo='Secretaria').update(nombre=f"Secretaria {nuevo_nombre}")
            else:
                responsables.filter(tipo='Secretaria').update(nombre=secretaria)

            if not secretaria_subrogante:
                responsables.filter(tipo='Secretaria Subrogante').update(nombre=f"Secretaria {nuevo_nombre}(s)")
            else:
                responsables.filter(tipo='Secretaria Subrogante').update(nombre=secretaria_subrogante)

            messages.success(request, 'Departamento modificado con éxito.')
            return redirect('home')
        messages.error(request, 'Error al modificar el departamento. Verifica los datos.')
    else:
        form = ModificarDepartamentoForm()
        departamentos = Departamento.objects.filter(activo=True)
        responsables_por_departamento = {}
        for dept in departamentos:
            responsables = dept.responsables.all()
            responsables_dict = {r.tipo: r.nombre for r in responsables}
            responsables_por_departamento[dept.nombre] = responsables_dict
        responsables_json = mark_safe(json.dumps(responsables_por_departamento))
        print("Departamentos disponibles en la vista modificar_departamento:", list(departamentos))
        print("Opciones del campo departamento:", form.fields['departamento'].choices)
        print("Responsables por departamento:", responsables_por_departamento)
    return render(request, 'accounts/modificar_departamento.html', {
        'form': form,
        'responsables_json': responsables_json
    })

@login_required
def eliminar_departamento(request):
    """Vista para deshabilitar un departamento"""
    limpiar_sesion_productos_salida(request)
    if not request.user.has_perm('accounts.can_manage_departments'):
        messages.error(request, 'No tienes permiso para deshabilitar departamentos.')
        return redirect('home')
    if request.method == 'POST':
        form = EliminarDepartamentoForm(request.POST)
        if form.is_valid():
            departamento_nombre = form.cleaned_data['departamento']
            departamento = Departamento.objects.get(nombre=departamento_nombre, activo=True)
            departamento.activo = False
            departamento.save()
            messages.success(request, 'Departamento deshabilitado con éxito.')
            return redirect('home')
        messages.error(request, 'Error al deshabilitar el departamento. Verifica los datos.')
    else:
        form = EliminarDepartamentoForm()
        departamentos = Departamento.objects.filter(activo=True)
        print("Departamentos disponibles en la vista eliminar_departamento:", list(departamentos))
        print("Opciones del campo departamento en la vista:", form.fields['departamento'].choices)
    return render(request, 'accounts/eliminar_departamento.html', {'form': form})

@login_required
def funcionarios_por_departamento(request):
    """Vista para obtener los responsables de un departamento"""
    departamento = request.GET.get('departamento', '')
    if not departamento:
        return JsonResponse({'error': 'Departamento no especificado'}, status=400)

    try:
        departamento_obj = Departamento.objects.get(nombre=departamento)
        responsables = departamento_obj.responsables.all()
        responsables_list = [{'id': r.id, 'nombre': r.nombre} for r in responsables]
        print(f"Responsables encontrados para {departamento}: {responsables_list}")
        return JsonResponse({'funcionarios': responsables_list})
    except Departamento.DoesNotExist:
        return JsonResponse({'error': 'Departamento no encontrado'}, status=404)

# Vistas para gestión de usuarios
@login_required
@permission_required('accounts.can_access_admin', raise_exception=True)
@permission_required('accounts.can_manage_users', raise_exception=True)
def listar_usuarios(request):
    """Vista para listar usuarios con búsqueda y paginación"""
    limpiar_sesion_productos_salida(request)
    
    usuarios = CustomUser.objects.all().order_by('rut')
    form = SearchUserForm(request.GET or None)

    query_rut = request.GET.get('rut', '')
    query_nombre = request.GET.get('nombre', '')
    query_rol = request.GET.get('rol', '')

    if request.GET:
        if form.is_valid():
            print(f"Formulario válido: True")
            print(f"Datos limpiados: {form.cleaned_data}")
            query_rut = form.cleaned_data['rut']
            query_nombre = form.cleaned_data['nombre']
            query_rol = form.cleaned_data['rol']

            if query_rut:
                usuarios = usuarios.filter(rut__icontains=query_rut)
            if query_nombre:
                usuarios = usuarios.filter(nombre__icontains=query_nombre)
            if query_rol:
                print(f"Filtrando por rol: {query_rol}")
                usuarios = usuarios.filter(groups__name=query_rol)
        else:
            print(f"Formulario válido: False")
            print(f"Errores del formulario: {form.errors.as_json()}")
    else:
        query_rut = ''
        query_nombre = ''
        query_rol = ''

    page_obj = paginar_resultados(request, usuarios)

    context = {
        'page_obj': page_obj,
        'form': form,
        'query_rut': query_rut,
        'query_nombre': query_nombre,
        'query_rol': query_rol,
    }
    return render(request, 'accounts/listar_usuarios.html', context)

@login_required
@permission_required('accounts.can_access_admin', raise_exception=True)
@permission_required('accounts.can_manage_users', raise_exception=True)
def agregar_usuario(request):
    """Vista para agregar un nuevo usuario"""
    limpiar_sesion_productos_salida(request)
    
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            usuario = form.save()
            messages.success(request, f'Usuario {usuario.rut} creado con éxito.')
            return redirect('listar-usuarios')
        messages.error(request, 'Error al crear el usuario. Verifica los datos.')
    else:
        form = CustomUserCreationForm()
    
    return render(request, 'accounts/agregar_usuario.html', {'form': form})

@login_required
@permission_required('accounts.can_access_admin', raise_exception=True)
@permission_required('accounts.can_manage_users', raise_exception=True)
def editar_usuario(request, rut):
    """Vista para editar un usuario existente"""
    limpiar_sesion_productos_salida(request)
    
    usuario = get_object_or_404(CustomUser, rut=rut)
    
    if request.method == 'POST':
        form = CustomUserEditForm(request.POST, instance=usuario)
        if form.is_valid():
            # Verificar si el usuario autenticado está intentando editarse a sí mismo
            if request.user.rut == usuario.rut:
                # Verificar si el usuario autenticado es un administrador
                if request.user.has_perm('accounts.can_manage_users'):
                    # Obtener el grupo actual del usuario (puede ser None si no tiene grupo asignado)
                    current_group = usuario.groups.first()
                    current_group_id = current_group.id if current_group else None
                    # Obtener el grupo seleccionado en el formulario
                    new_group = form.cleaned_data['grupo']
                    # Asegurarse de que se haya seleccionado un grupo
                    if not new_group:
                        messages.error(request, 'Debe seleccionar un rol válido.')
                        return render(request, 'accounts/editar_usuario.html', {'form': form, 'usuario': usuario})
                    new_group_id = new_group.id
                    # Comparar si el grupo ha cambiado
                    if current_group_id != new_group_id:
                        error_message = f'El usuario {request.user.rut} intentó cambiar su propio rol de {current_group.name if current_group else "Ninguno"} a {new_group.name}.'
                        logger.info(error_message)
                        messages.error(request, 'No puedes cambiar tu propio rol. Pide a otro administrador que realice este cambio.')
                        return render(request, 'accounts/editar_usuario.html', {'form': form, 'usuario': usuario})
            
            # Si no hay restricción, proceder con la actualización
            usuario = form.save()
            password = form.cleaned_data.get('password')
            if password:
                usuario.set_password(password)
                usuario.save()
                messages.info(request, 'La contraseña ha sido actualizada.')
            messages.success(request, f'Usuario {usuario.rut} actualizado con éxito.')
            return redirect('listar-usuarios')
        messages.error(request, 'Error al actualizar el usuario. Verifica los datos.')
    else:
        form = CustomUserEditForm(instance=usuario)
    
    return render(request, 'accounts/editar_usuario.html', {'form': form, 'usuario': usuario})

@login_required
@permission_required('accounts.can_access_admin', raise_exception=True)
@permission_required('accounts.can_manage_users', raise_exception=True)
def deshabilitar_usuario(request, rut):
    """Vista para deshabilitar o habilitar un usuario"""
    limpiar_sesion_productos_salida(request)
    
    usuario = get_object_or_404(CustomUser, rut=rut)
    
    if request.method == 'POST':
        if usuario.is_active:
            usuario.is_active = False
            usuario.save()
            messages.success(request, f'Usuario {usuario.rut} deshabilitado con éxito.')
        else:
            usuario.is_active = True
            usuario.save()
            messages.success(request, f'Usuario {usuario.rut} habilitado con éxito.')
        return redirect('listar-usuarios')
    
    return render(request, 'accounts/deshabilitar_usuario.html', {'usuario': usuario})

@login_required
@permission_required('accounts.can_access_admin', raise_exception=True)
def verify_password(request):
    """Vista para verificar la contraseña del administrador antes de acceder al panel de administración"""
    if not request.user.is_staff or not request.user.is_superuser:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'No tienes permisos suficientes para acceder al panel de administración.'}, status=403)
        messages.error(request, 'No tienes permisos suficientes para acceder al panel de administración. Contacta a un administrador.')
        return redirect('home')

    if request.method == 'POST':
        password = request.POST.get('password')
        user = request.user

        if user.check_password(password):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            messages.success(request, 'Contraseña verificada correctamente. Redirigiendo al panel de administración...')
            return redirect('/admin/')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': 'Contraseña incorrecta. Por favor, intenta de nuevo.'})
            messages.error(request, 'Contraseña incorrecta. Por favor, intenta de nuevo.')
            return render(request, 'accounts/verify_password.html')

    return render(request, 'accounts/verify_password.html')

# Vistas para la gestión de categorías
@login_required
@permission_required('accounts.can_edit', raise_exception=True)
def agregar_categoria(request):
    """Vista para agregar una nueva categoría"""
    limpiar_sesion_productos_salida(request)
    if not request.user.has_perm('accounts.can_edit'):

        messages.error(request, 'No tienes permiso para agregar categorías.')
        return redirect('home')

    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            categoria = form.save(commit=False)
            categoria.activo = True  # Asegurar que la nueva categoría esté activa
            categoria.save()
            messages.success(request, f'La categoría "{categoria.nombre}" ha sido agregada con éxito.')
            return redirect('listar-productos')  # Redirigir a la lista de productos
        else:
            messages.error(request, 'Error al agregar la categoría. Verifica los datos.')
    else:
        form = CategoriaForm()  

    return render(request, 'accounts/agregar_categoria.html', {'form': form})

@login_required
@permission_required('accounts.can_edit', raise_exception=True)
def modificar_categoria(request):
    """Vista para modificar una categoría existente"""
    limpiar_sesion_productos_salida(request)
    if not request.user.has_perm('accounts.can_edit'):
        messages.error(request, 'No tienes permiso para modificar categorías.')
        return redirect('home')

    if request.method == 'POST':
        form = ModificarCategoriaForm(request.POST)
        if form.is_valid():
            categoria_nombre = form.cleaned_data['categoria']
            nuevo_nombre = form.cleaned_data['nuevo_nombre']
            try:
                categoria = Categoria.objects.get(nombre=categoria_nombre, activo=True)
                categoria.nombre = nuevo_nombre
                categoria.save()
                messages.success(request, f'La categoría ha sido renombrada a "{nuevo_nombre}" con éxito.')
                return redirect('listar-productos')  # Redirigir a la lista de productos
            except Categoria.DoesNotExist:
                messages.error(request, 'La categoría seleccionada no existe.')
        else:
            messages.error(request, 'Error al modificar la categoría. Verifica los datos.')
    else:
        form = ModificarCategoriaForm()

    return render(request, 'accounts/modificar_categoria.html', {'form': form})

@login_required
@permission_required('accounts.can_edit', raise_exception=True)
def eliminar_categoria(request):
    """Vista para deshabilitar una categoría existente"""
    limpiar_sesion_productos_salida(request)
    if not request.user.has_perm('accounts.can_edit'):
        messages.error(request, 'No tienes permiso para deshabilitar categorías.')
        return redirect('home')

    if request.method == 'POST':
        form = EliminarCategoriaForm(request.POST)
        if form.is_valid():
            categoria_nombre = form.cleaned_data['categoria']
            try:
                categoria = Categoria.objects.get(nombre=categoria_nombre, activo=True)
                # Verificar si hay productos asociados a esta categoría
                if Producto.objects.filter(categoria=categoria).exists():
                    messages.error(request, 'No se puede deshabilitar esta categoría porque tiene productos asociados.')
                    return redirect('eliminar-categoria')
                categoria.activo = False
                categoria.save()
                messages.success(request, f'La categoría "{categoria.nombre}" ha sido deshabilitada con éxito.')
                return redirect('listar-productos')  # Redirigir a la lista de productos
            except Categoria.DoesNotExist:
                messages.error(request, 'La categoría seleccionada no existe.')
        else:
            messages.error(request, 'Error al deshabilitar la categoría. Verifica los datos.')
    else:
        form = EliminarCategoriaForm()

    return render(request, 'accounts/eliminar_categoria.html', {'form': form})

@login_required
def control_vencimientos(request):
    """Vista para el control de vencimientos de productos con información de lotes."""
    from datetime import date, timedelta
    hoy = date.today()
    
    # Obtener productos con vencimiento que tienen stock
    productos_base = Producto.objects.filter(
        tiene_vencimiento=True,
        stock__gt=0
    ).select_related('categoria').prefetch_related('lotes')
    
    # Filtrado por estado
    estado_filtro = request.GET.get('estado', 'todos')
    busqueda = request.GET.get('busqueda', '')
    
    # Crear lista de productos con información de lotes para filtrado
    productos_info = []
    for producto in productos_base:
        # Usar el estado de vencimiento completo que considera todos los lotes
        estado_vencimiento = producto.get_estado_vencimiento_completo()
        proximo_vencimiento = producto.get_proximo_vencimiento()
        lotes_detalle = producto.get_lotes_detalle()
        
        # Calcular días restantes del lote más próximo a vencer
        dias_restantes = None
        if proximo_vencimiento:
            dias_restantes = (proximo_vencimiento - hoy).days
        
        productos_info.append({
            'producto': producto,
            'estado_vencimiento': estado_vencimiento,
            'proximo_vencimiento': proximo_vencimiento,
            'dias_restantes': dias_restantes,
            'lotes_detalle': lotes_detalle,
            'total_lotes': len(lotes_detalle)
        })
    
    # Filtrar por estado
    if estado_filtro == 'vencidos':
        productos_info = [p for p in productos_info if p['estado_vencimiento'] == 'Vencido']
    elif estado_filtro == 'criticos':
        productos_info = [p for p in productos_info if p['estado_vencimiento'] in ['Vence Hoy', 'Crítico']]
    elif estado_filtro == 'precaucion':
        productos_info = [p for p in productos_info if p['estado_vencimiento'] == 'Precaución']
    
    # Búsqueda por código o descripción
    if busqueda:
        productos_info = [
            p for p in productos_info 
            if (busqueda.lower() in p['producto'].codigo_barra.lower() or 
                busqueda.lower() in p['producto'].descripcion.lower())
        ]
    
    # Ordenar por días restantes (más críticos primero)
    productos_info.sort(key=lambda x: (
        0 if x['dias_restantes'] is None or x['dias_restantes'] < 0 else x['dias_restantes'],
        x['producto'].descripcion
    ))
    
    # Paginación con numeración dinámica
    productos_paginados = paginar_resultados_dinamico(request, productos_info, 10)
    
    # Calcular estadísticas usando el estado completo de lotes
    estadisticas = {'vencidos': 0, 'criticos': 0, 'precaucion': 0, 'normal': 0}
    for producto in productos_base:
        estado = producto.get_estado_vencimiento_completo()
        if estado == 'Vencido':
            estadisticas['vencidos'] += 1
        elif estado in ['Vence Hoy', 'Crítico']:
            estadisticas['criticos'] += 1
        elif estado == 'Precaución':
            estadisticas['precaucion'] += 1
        else:
            estadisticas['normal'] += 1
    
    context = {
        'productos': productos_paginados,
        'estado_filtro': estado_filtro,
        'busqueda': busqueda,
        'productos_vencidos': estadisticas['vencidos'],
        'productos_criticos': estadisticas['criticos'],
        'productos_precaucion': estadisticas['precaucion'],
        'productos_normal': estadisticas['normal'],
        'total_con_vencimiento': len(productos_info),
        'hoy': hoy,
        'mostrar_lotes': True,  # Flag para mostrar información de lotes en template
    }
    
    return render(request, 'accounts/control_vencimientos.html', context)

@login_required
def exportar_vencimientos_excel(request):
    """Exporta el control de vencimientos a Excel."""
    from datetime import date, timedelta
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    hoy = date.today()
    # Obtener productos con vencimiento y stock > 0
    productos = Producto.objects.filter(
        tiene_vencimiento=True
    ).select_related('categoria').prefetch_related('lotes').order_by('descripcion')

    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control de Vencimientos"

    # Encabezados principales
    headers = [
        'Código de Barra', 'Descripción', 'Categoría', 'Stock Total',
        'N° Lote', 'Stock Lote', 'Fecha Vencimiento Lote', 'Días Restantes', 'Estado Lote', 'Estado Producto'
    ]

    # Estilo para encabezados
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1a3c5e', end_color='1a3c5e', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    row = 2
    for producto in productos:
        lotes = list(producto.lotes.filter(stock__gt=0).order_by('fecha_vencimiento'))
        estado_producto = producto.get_estado_vencimiento_completo()
        # Si tiene lotes con stock, mostrar cada lote
        if lotes:
            for lote in lotes:
                ws.cell(row=row, column=1, value=producto.codigo_barra)
                ws.cell(row=row, column=2, value=producto.descripcion)
                ws.cell(row=row, column=3, value=producto.categoria.nombre if producto.categoria else 'Sin categoría')
                ws.cell(row=row, column=4, value=producto.stock)
                ws.cell(row=row, column=5, value=lote.numero_lote)
                ws.cell(row=row, column=6, value=lote.stock)
                ws.cell(row=row, column=7, value=lote.fecha_vencimiento.strftime('%d/%m/%Y'))
                ws.cell(row=row, column=8, value=lote.get_dias_para_vencer())
                ws.cell(row=row, column=9, value=lote.get_estado_vencimiento())
                ws.cell(row=row, column=10, value=estado_producto)
                # Colorear según el estado del lote
                color_fill = None
                estado_lote = lote.get_estado_vencimiento()
                if estado_lote == 'Vencido':
                    color_fill = PatternFill(start_color='ffebee', end_color='ffebee', fill_type='solid')
                elif estado_lote in ['Vence Hoy', 'Crítico']:
                    color_fill = PatternFill(start_color='fff3e0', end_color='fff3e0', fill_type='solid')
                elif estado_lote == 'Precaución':
                    color_fill = PatternFill(start_color='e8f5e8', end_color='e8f5e8', fill_type='solid')
                if color_fill:
                    for col in range(1, 11):
                        ws.cell(row=row, column=col).fill = color_fill
                row += 1
        else:
            # Producto sin lotes activos pero con vencimiento
            ws.cell(row=row, column=1, value=producto.codigo_barra)
            ws.cell(row=row, column=2, value=producto.descripcion)
            ws.cell(row=row, column=3, value=producto.categoria.nombre if producto.categoria else 'Sin categoría')
            ws.cell(row=row, column=4, value=producto.stock)
            ws.cell(row=row, column=5, value='-')
            ws.cell(row=row, column=6, value='-')
            fecha_venc = producto.fecha_vencimiento.strftime('%d/%m/%Y') if producto.fecha_vencimiento else '-'
            ws.cell(row=row, column=7, value=fecha_venc)
            dias_rest = producto.get_dias_para_vencer() if producto.fecha_vencimiento else '-'
            ws.cell(row=row, column=8, value=dias_rest)
            ws.cell(row=row, column=9, value=producto.get_estado_vencimiento())
            ws.cell(row=row, column=10, value=estado_producto)
            # Colorear según el estado del producto
            color_fill = None
            estado = producto.get_estado_vencimiento()
            if estado == 'Vencido':
                color_fill = PatternFill(start_color='ffebee', end_color='ffebee', fill_type='solid')
            elif estado in ['Vence Hoy', 'Crítico']:
                color_fill = PatternFill(start_color='fff3e0', end_color='fff3e0', fill_type='solid')
            elif estado == 'Precaución':
                color_fill = PatternFill(start_color='e8f5e8', end_color='e8f5e8', fill_type='solid')
            if color_fill:
                for col in range(1, 11):
                    ws.cell(row=row, column=col).fill = color_fill
            row += 1

    # Ajustar ancho de columnas
    column_widths = [15, 40, 20, 12, 10, 12, 18, 15, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="control_vencimientos_{hoy.strftime('%Y%m%d')}.xlsx"'

    wb.save(response)
    return response

@login_required 
def agregar_vencimiento_producto(request):
    """Vista para agregar y modificar fechas de vencimiento de productos y lotes."""
    if not request.user.has_perm('accounts.can_edit'):
        messages.error(request, 'No tienes permiso para agregar o modificar vencimientos.')
        return redirect('home')
    
    # Obtener todos los productos para mostrar en la vista
    productos = Producto.objects.all().select_related('categoria').prefetch_related('lotes').order_by('descripcion')
    
    # Filtros
    query_codigo = request.GET.get('codigo_barra', '').strip()
    query_descripcion = request.GET.get('descripcion', '').strip()
    query_categoria = request.GET.get('categoria', '').strip()
    tipo_filtro = request.GET.get('tipo', 'todos')  # todos, sin_vencimiento, con_vencimiento
    
    # Aplicar filtros
    if query_codigo:
        productos = productos.filter(codigo_barra__icontains=query_codigo)
    if query_descripcion:
        productos = productos.filter(descripcion__icontains=query_descripcion)
    if query_categoria:
        productos = productos.filter(categoria__nombre__icontains=query_categoria)
    
    # Filtrar por tipo de vencimiento
    if tipo_filtro == 'sin_vencimiento':
        productos = productos.filter(tiene_vencimiento=False)
    elif tipo_filtro == 'con_vencimiento':
        productos = productos.filter(tiene_vencimiento=True)
    
    # Preparar información adicional para cada producto
    productos_info = []
    for producto in productos:
        info = {
            'producto': producto,
            'lotes_detalle': [],
            'total_lotes': 0,
            'proximo_vencimiento': None,
            'estado_vencimiento': 'Sin Vencimiento'
        }
        
        if producto.tiene_vencimiento:
            info['lotes_detalle'] = producto.get_lotes_detalle()
            info['total_lotes'] = len(info['lotes_detalle'])
            info['proximo_vencimiento'] = producto.get_proximo_vencimiento()
            info['estado_vencimiento'] = producto.get_estado_vencimiento_completo()
        
        productos_info.append(info)
    
    # Paginación
    page_obj = paginar_resultados(request, productos_info, 20)
    
    # Obtener categorías para el filtro
    from .models import Categoria
    categorias = Categoria.objects.filter(activo=True).order_by('nombre')
    
    context = {
        'page_obj': page_obj,
        'query_codigo': query_codigo,
        'query_descripcion': query_descripcion,
        'query_categoria': query_categoria,
        'tipo_filtro': tipo_filtro,
        'categorias': categorias,
        'total_productos': productos.count(),
        'productos_sin_vencimiento': productos.filter(tiene_vencimiento=False).count(),
        'productos_con_vencimiento': productos.filter(tiene_vencimiento=True).count(),
    }
    
    return render(request, 'accounts/agregar_vencimiento.html', context)

@login_required
def agregar_vencimiento_ajax(request):
    """Vista AJAX para agregar vencimiento a un producto sin fecha de vencimiento."""
    if not request.user.has_perm('accounts.can_edit'):
        return JsonResponse({'success': False, 'error': 'No tienes permiso para realizar esta acción.'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'})
    
    try:
        codigo_barra = request.POST.get('codigo_barra')
        fecha_vencimiento = request.POST.get('fecha_vencimiento')
        
        if not codigo_barra or not fecha_vencimiento:
            return JsonResponse({'success': False, 'error': 'Faltan datos requeridos.'})
        
        # Validar el producto
        try:
            producto = Producto.objects.get(codigo_barra=codigo_barra)
        except Producto.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Producto no encontrado.'})
        
        if producto.tiene_vencimiento:
            return JsonResponse({'success': False, 'error': 'Este producto ya tiene fecha de vencimiento.'})
        
        # Validar la fecha
        from datetime import datetime, date
        try:
            fecha_obj = datetime.strptime(fecha_vencimiento, '%Y-%m-%d').date()
            if fecha_obj <= date.today():
                return JsonResponse({'success': False, 'error': 'La fecha de vencimiento debe ser posterior a hoy.'})
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Formato de fecha inválido.'})
        
        # Activar vencimiento en el producto
        producto.tiene_vencimiento = True
        producto.fecha_vencimiento = fecha_obj
        producto.save()
        
        # Si el producto tiene stock, crear un lote inicial
        if producto.stock > 0:
            stock_inicial = producto.stock  # Guardar el stock actual
            lote = LoteProducto.objects.create(
                producto=producto,
                numero_lote=producto.get_proximo_numero_lote(),
                fecha_vencimiento=fecha_obj,
                stock=stock_inicial
            )
            # No cambiar el stock del producto, ya está correcto
            # Solo sincronizar para asegurar consistencia
            producto.sincronizar_stock_con_lotes()
            
            return JsonResponse({
                'success': True, 
                'message': f'Vencimiento agregado exitosamente. Se creó el Lote #{lote.numero_lote} con {lote.stock} unidades.',
                'lote_creado': True,
                'numero_lote': lote.numero_lote
            })
        else:
            return JsonResponse({
                'success': True, 
                'message': 'Vencimiento agregado exitosamente. El producto ahora puede recibir lotes con fechas de vencimiento.',
                'lote_creado': False
            })
            
    except Exception as e:
        logger.error(f"Error al agregar vencimiento: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Error interno: {str(e)}'})

@login_required
def modificar_vencimiento_producto_ajax(request):
    """Vista AJAX para modificar la fecha de vencimiento principal de un producto."""
    if not request.user.has_perm('accounts.can_edit'):
        return JsonResponse({'success': False, 'error': 'No tienes permiso para realizar esta acción.'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'})
    
    try:
        codigo_barra = request.POST.get('codigo_barra')
        nueva_fecha = request.POST.get('nueva_fecha')
        
        if not codigo_barra or not nueva_fecha:
            return JsonResponse({'success': False, 'error': 'Faltan datos requeridos.'})
        
        # Validar el producto
        try:
            producto = Producto.objects.get(codigo_barra=codigo_barra)
        except Producto.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Producto no encontrado.'})
        
        if not producto.tiene_vencimiento:
            return JsonResponse({'success': False, 'error': 'Este producto no tiene fecha de vencimiento.'})
        
        # Validar la fecha
        from datetime import datetime, date
        try:
            fecha_obj = datetime.strptime(nueva_fecha, '%Y-%m-%d').date()
            if fecha_obj <= date.today():
                return JsonResponse({'success': False, 'error': 'La fecha de vencimiento debe ser posterior a hoy.'})
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Formato de fecha inválido.'})
        
        # Actualizar la fecha principal del producto
        fecha_anterior = producto.fecha_vencimiento
        producto.fecha_vencimiento = fecha_obj
        producto.save()
        
        # CORRECCIÓN: También actualizar TODOS los lotes del producto
        lotes_actualizados = 0
        if producto.lotes.exists():
            lotes_actualizados = producto.lotes.update(fecha_vencimiento=fecha_obj)
        
        return JsonResponse({
            'success': True, 
            'message': f'Fecha de vencimiento actualizada de {fecha_anterior.strftime("%d/%m/%Y") if fecha_anterior else "Sin fecha"} a {fecha_obj.strftime("%d/%m/%Y")}. {lotes_actualizados} lotes actualizados.',
            'nueva_fecha': fecha_obj.strftime("%Y-%m-%d"),
            'lotes_actualizados': lotes_actualizados
        })
            
    except Exception as e:
        logger.error(f"Error al modificar vencimiento del producto: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Error interno: {str(e)}'})

@login_required
def modificar_vencimiento_lote_ajax(request):
    """Vista AJAX para modificar la fecha de vencimiento de un lote específico."""
    if not request.user.has_perm('accounts.can_edit'):
        return JsonResponse({'success': False, 'error': 'No tienes permiso para realizar esta acción.'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'})
    
    try:
        codigo_barra = request.POST.get('codigo_barra')
        numero_lote = request.POST.get('numero_lote')
        nueva_fecha = request.POST.get('nueva_fecha')
        
        if not codigo_barra or not numero_lote or not nueva_fecha:
            return JsonResponse({'success': False, 'error': 'Faltan datos requeridos.'})
        
        # Validar el producto
        try:
            producto = Producto.objects.get(codigo_barra=codigo_barra)
        except Producto.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Producto no encontrado.'})
        
        # Validar el lote
        try:
            lote = LoteProducto.objects.get(producto=producto, numero_lote=int(numero_lote))
        except (LoteProducto.DoesNotExist, ValueError):
            return JsonResponse({'success': False, 'error': 'Lote no encontrado.'})
        
        # Validar la fecha
        from datetime import datetime, date
        try:
            fecha_obj = datetime.strptime(nueva_fecha, '%Y-%m-%d').date()
            if fecha_obj <= date.today():
                return JsonResponse({'success': False, 'error': 'La fecha de vencimiento debe ser posterior a hoy.'})
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Formato de fecha inválido.'})
        
        # Actualizar la fecha del lote
        fecha_anterior = lote.fecha_vencimiento
        lote.fecha_vencimiento = fecha_obj
        lote.save()
        
        return JsonResponse({
            'success': True, 
            'message': f'Fecha del Lote #{lote.numero_lote} actualizada de {fecha_anterior.strftime("%d/%m/%Y")} a {fecha_obj.strftime("%d/%m/%Y")}.',
            'nuevo_estado': lote.get_estado_vencimiento(),
            'nuevos_dias': lote.get_dias_para_vencer(),
            'nuevo_color': lote.get_color_estado_vencimiento()
        })
            
    except Exception as e:
        logger.error(f"Error al modificar vencimiento del lote: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Error interno: {str(e)}'})

@login_required
def obtener_lotes_producto_ajax(request):
    """Vista AJAX para obtener los lotes de un producto."""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'})
    
    try:
        codigo_barra = request.GET.get('codigo_barra')
        
        if not codigo_barra:
            return JsonResponse({'success': False, 'error': 'Código de barra requerido.'})
        
        # Validar el producto
        try:
            producto = Producto.objects.get(codigo_barra=codigo_barra)
        except Producto.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Producto no encontrado.'})
        
        if not producto.tiene_vencimiento:
            return JsonResponse({'success': False, 'error': 'Este producto no maneja lotes.'})
        
        # Obtener lotes con stock
        lotes_detalle = producto.get_lotes_detalle()
        
        lotes_data = []
        for lote in lotes_detalle:
            lotes_data.append({
                'numero_lote': lote['numero_lote'],
                'fecha_vencimiento': lote['fecha_vencimiento'].strftime('%Y-%m-%d'),
                'fecha_vencimiento_display': lote['fecha_vencimiento'].strftime('%d/%m/%Y'),
                'stock': lote['stock'],
                'dias_restantes': lote['dias_restantes'],
                'estado': lote['estado'],
                'color': lote['color']
            })
        
        return JsonResponse({
            'success': True, 
            'lotes': lotes_data,
            'total_lotes': len(lotes_data),
            'producto': {
                'descripcion': producto.descripcion,
                'fecha_vencimiento': producto.fecha_vencimiento.strftime('%Y-%m-%d') if producto.fecha_vencimiento else '',
                'fecha_vencimiento_display': producto.fecha_vencimiento.strftime('%d/%m/%Y') if producto.fecha_vencimiento else 'Sin fecha'
            }
        })
            
    except Exception as e:
        logger.error(f"Error al obtener lotes del producto: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Error interno: {str(e)}'})

@login_required
def obtener_datos_producto_ajax(request):
    """Vista AJAX para obtener datos actualizados de un producto."""
    if request.method == 'GET':
        try:
            codigo_barra = request.GET.get('codigo_barra')
            if not codigo_barra:
                return JsonResponse({'success': False, 'error': 'Código de barra requerido.'})
            
            producto = Producto.objects.get(codigo_barra=codigo_barra)
            
            # Calcular datos actualizados
            estado_vencimiento = producto.get_estado_vencimiento_completo()
            proximo_vencimiento = producto.get_proximo_vencimiento()
            total_lotes = producto.lotes.filter(stock__gt=0).count()
            
            data = {
                'success': True,
                'producto': {
                    'codigo_barra': producto.codigo_barra,
                    'descripcion': producto.descripcion,
                    'estado_vencimiento': estado_vencimiento,
                    'proximo_vencimiento': proximo_vencimiento.strftime('%Y-%m-%d') if proximo_vencimiento else None,
                    'proximo_vencimiento_display': proximo_vencimiento.strftime('%d/%m/%Y') if proximo_vencimiento else None,
                    'total_lotes': total_lotes,
                    'tiene_vencimiento': producto.tiene_vencimiento
                }
            }
            
            return JsonResponse(data)
            
        except Producto.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Producto no encontrado.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error interno: {str(e)}'})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'})